"""Plan-Excel generator (v0.3).

Veranderingen t.o.v. v0.2:
- Tactiek_id bevat geen productlijn meer (BON/SPR eruit).
- Single-flight mode: geen flight-kolommen, geen Flights-tab, geen D-prefix.
- Multi-flight mode: kolom A = flight_nr, en tussen flight-groepen een
  banner-rij ("Flight 1 - Launch  periode  EUR X") voor visuele grouping.
- Plan-data-tabel wordt nu handmatig via openpyxl geschreven zodat die
  banner-rijen netjes tussen de groepen passen. De evaluatiemachine kan
  ze makkelijk skippen (banner-rijen hebben enkel col A, geen tactiek_id).
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PLAN_COLUMNS = [
    # Flight-kolommen eerst (worden gedropt bij single-flight mode)
    "flight_nr",
    "flight_naam",
    "tactiek_id",
    "doelstelling",
    "kanaal",
    "kanaal_type",
    "subkanaal",
    "format",
    "doelgroep",
    "creatief",
    "flight_start",
    "flight_eind",
    "budget",
    # Kostenmetrieken - inputs van de planner (of default uit cfg.kpi_targets)
    "cpm",
    "cpc",
    "cpa",
    "ctr_pct",
    "gcf",
    # Volume-metrieken - afgeleid uit budget + kostenmetrieken
    "impressies",
    "clicks",
    "bereik",
    "conversies",
    "ecpc",         # effective cost per click - afgeleid = budget / clicks
    "kpi_primary",
    "remarks",
]


# -----------------------------------------------------------------------------
#  Uren / header
# -----------------------------------------------------------------------------

@dataclass
class UrenPost:
    categorie: str
    tarief: float
    aantal_uren: float

    @property
    def kosten(self) -> float:
        return round(self.tarief * self.aantal_uren, 2)


def uren_from_klantconfig(klant_cfg: dict) -> list:
    """Tarief default = €110 voor alle klanten (aanpasbaar in UI).
    Posten: uit klantconfig.uren_defaults.posten, anders generieke fallback."""
    ud = (klant_cfg or {}).get("uren_defaults") or {}
    tarief = float(ud.get("tarief", 110))
    posten = ud.get("posten") or {
        "setup campagne":    8,
        "beheer campagne":   12,
        "reporting campagne": 12,
        "project management": 6,
    }
    return [UrenPost(str(k), tarief, float(v)) for k, v in posten.items()]


@dataclass
class PlanHeader:
    klant: str
    campagne: str
    start: date
    eind: date
    totaal_incl_btw: float
    btw_pct: float = 21.0
    uren_posten: list = field(default_factory=list)

    @property
    def totaal_ex_btw(self) -> float:
        f = 1.0 + (self.btw_pct / 100.0)
        return round(self.totaal_incl_btw / f, 2) if f else round(self.totaal_incl_btw, 2)

    @property
    def uren_ex_btw(self) -> float:
        return round(sum(p.kosten for p in self.uren_posten), 2)

    @property
    def media_ex_btw(self) -> float:
        return round(self.totaal_ex_btw - self.uren_ex_btw, 2)

    @property
    def totaal_uren(self) -> float:
        return round(sum(p.aantal_uren for p in self.uren_posten), 2)


# -----------------------------------------------------------------------------
#  Plan-context + row-builder
# -----------------------------------------------------------------------------

@dataclass
class PlanContext:
    klant_cfg: dict
    campagne: dict
    auteur: str = "planningsmachine"
    versie: str = "0.3"
    benchmarks_used: dict = field(default_factory=dict)
    keybeliefs_used: list = field(default_factory=list)
    header: Any = None
    # Verwachte overlap tussen kanalen (fractie 0..1). Klein plan ~0.03, medium
    # 0.12-0.19, groot plan tot 0.30. Wordt gebruikt in totalen-rij om het
    # samengetelde bereik te ontdubbelen: netto = SUM * (1 - overlap).
    overlap_factor: float = 0.08


def _k(cfg: dict, *path, default=None):
    cur = cfg
    for p in path:
        if not isinstance(cur, dict):
            return default
        cur = cur.get(p, default)
    return cur


def _kanaal_type(cfg: dict, kanaal: str) -> str:
    k = (cfg.get("kanalen") or {}).get(kanaal) or {}
    return k.get("type", "") if isinstance(k, dict) else ""


def _kpi_defaults(cfg: dict, fase: str) -> dict:
    return (cfg.get("kpi_targets") or {}).get(fase) or {}


def _first_target(kpi_cfg: dict):
    for k, v in (kpi_cfg or {}).items():
        if k.startswith("target_"):
            return v
    return None


def _iso(d) -> str:
    if d is None or d == "":
        return ""
    if isinstance(d, (date, datetime)):
        return d.strftime("%Y-%m-%d")
    return str(d)[:10]


def _month_of(d):
    if not d:
        return None
    if isinstance(d, (date, datetime)):
        return d.month
    try:
        return date.fromisoformat(str(d)[:10]).month
    except Exception:
        return None


def build_plan_rows(tactieken: list, ctx: PlanContext) -> pd.DataFrame:
    cfg = ctx.klant_cfg

    rows = []
    for t in tactieken:
        fase = t.get("fase", "")
        kanaal = t.get("kanaal", "")
        media = float(t.get("budget", 0) or 0)
        kpi = _kpi_defaults(cfg, fase)
        # Kostenmetrieken: input van tactiek of default uit fase-kpi_targets
        cpm     = t.get("cpm")     if t.get("cpm")     is not None else kpi.get("cpm")
        cpc     = t.get("cpc")     if t.get("cpc")     is not None else kpi.get("cpc")
        cpa     = t.get("cpa")     if t.get("cpa")     is not None else kpi.get("cpa")
        ctr_pct = t.get("ctr_pct") if t.get("ctr_pct") is not None else kpi.get("ctr_pct")
        gcf     = t.get("gcf")     if t.get("gcf")     is not None else kpi.get("gcf")

        # Volume-metrieken afgeleid
        impressies = round(media / cpm * 1000) if cpm and cpm > 0 else None
        if cpc and cpc > 0:
            clicks = round(media / cpc)
        elif impressies and ctr_pct:
            clicks = round(impressies * ctr_pct / 100)
        else:
            clicks = None
        bereik     = round(impressies / gcf) if (impressies and gcf and gcf > 0) else None
        conversies = round(media / cpa)      if (cpa and cpa > 0) else None
        # eCPC = budget / clicks (alleen zinvol als clicks > 0)
        ecpc       = round(media / clicks, 2) if (clicks and clicks > 0) else None

        rows.append({
            "flight_nr": t.get("flight_nr"),
            "flight_naam": t.get("flight_naam", ""),
            "tactiek_id": t.get("tactiek_id", ""),
            "doelstelling": fase,
            "kanaal": kanaal,
            "kanaal_type": _kanaal_type(cfg, kanaal),
            "subkanaal": t.get("subkanaal", ""),
            "format": t.get("format", ""),
            "doelgroep": t.get("doelgroep", ""),
            "creatief": t.get("creatief", ""),
            "flight_start": _iso(t.get("flight_start")),
            "flight_eind": _iso(t.get("flight_eind")),
            "budget": round(media, 2),
            "cpm": cpm,
            "cpc": cpc,
            "cpa": cpa,
            "ctr_pct": ctr_pct,
            "gcf": gcf,
            "impressies": impressies,
            "clicks": clicks,
            "bereik": bereik,
            "conversies": conversies,
            "ecpc": ecpc,
            "kpi_primary": t.get("kpi_primary") or kpi.get("primary", ""),
            "remarks": t.get("remarks", ""),
        })
    return pd.DataFrame(rows, columns=PLAN_COLUMNS)


# -----------------------------------------------------------------------------
#  Filters
# -----------------------------------------------------------------------------

def _hide_zero_rows(df: pd.DataFrame, value_col: str) -> pd.DataFrame:
    if df is None or value_col not in df.columns:
        return df
    mask = pd.to_numeric(df[value_col], errors="coerce").fillna(0) > 0
    return df.loc[mask].reset_index(drop=True)


def _is_multi_flight(df: pd.DataFrame) -> bool:
    if df is None or "flight_nr" not in df.columns:
        return False
    distinct = pd.to_numeric(df["flight_nr"], errors="coerce").dropna().unique()
    return len(distinct) > 1


# -----------------------------------------------------------------------------
#  Styling
# -----------------------------------------------------------------------------

_BOLD = Font(bold=True)
_BOLD_W = Font(bold=True, color="FFFFFF")
_BOLD_BLUE = Font(bold=True, color="1F3864")
_FILL_DARK = PatternFill("solid", fgColor="305496")
_FILL_MED  = PatternFill("solid", fgColor="8EA9DB")
_FILL_LIGHT = PatternFill("solid", fgColor="D9E1F2")
_FILL_YELLOW = PatternFill("solid", fgColor="FFF2CC")     # totalen-rij basis
_FILL_ORANGE = PatternFill("solid", fgColor="F4B084")     # gewogen-gem cellen
_THIN = Side(style="thin", color="808080")
_BOX = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_EUR = '#,##0;[Red]-#,##0'
_DATE = "dd-mm-yyyy"


# -----------------------------------------------------------------------------
#  Plan-header-kader (rijen 1-8)
# -----------------------------------------------------------------------------

def _write_plan_header(ws, h: PlanHeader) -> int:
    left = [
        ("klant",                   h.klant),
        ("campagne naam",           h.campagne),
        ("totaal budget ex btw",    h.totaal_ex_btw),
        ("totaal budget incl btw",  h.totaal_incl_btw),
        ("uren ex btw",             h.uren_ex_btw),
        ("media ex btw",            h.media_ex_btw),
        ("start datum",             h.start),
        ("eind datum",              h.eind),
    ]
    for i, (label, val) in enumerate(left, start=1):
        c_lbl = ws.cell(row=i, column=1, value=label)
        c_val = ws.cell(row=i, column=2, value=val)
        c_lbl.fill = _FILL_LIGHT; c_lbl.font = _BOLD; c_lbl.border = _BOX
        c_val.border = _BOX
        if isinstance(val, (int, float)) and not isinstance(val, bool):
            c_val.number_format = _EUR
        if isinstance(val, (date, datetime)):
            c_val.number_format = _DATE

    hdr = [("Uren", 4), ("tarief", 5), ("aantal uren", 6), ("kosten", 7)]
    for label, col in hdr:
        c = ws.cell(row=1, column=col, value=label)
        c.fill = _FILL_DARK; c.font = _BOLD_W; c.border = _BOX
        c.alignment = Alignment(horizontal="left" if col == 4 else "right")

    for i, p in enumerate(h.uren_posten, start=2):
        for c, val in zip((4, 5, 6, 7), (p.categorie, p.tarief, p.aantal_uren, p.kosten)):
            cell = ws.cell(row=i, column=c, value=val)
            cell.border = _BOX
            if c in (5, 7):
                cell.number_format = _EUR

    total_row = 1 + len(h.uren_posten) + 2
    c_lbl = ws.cell(row=total_row, column=4, value="totaal")
    c_uur = ws.cell(row=total_row, column=6, value=h.totaal_uren)
    c_kos = ws.cell(row=total_row, column=7, value=h.uren_ex_btw)
    for c in (c_lbl, c_uur, c_kos):
        c.font = _BOLD; c.border = _BOX
    c_kos.number_format = _EUR

    for col, width in [("A", 24), ("B", 18), ("D", 22), ("E", 10), ("F", 13), ("G", 12)]:
        ws.column_dimensions[col].width = width

    return max(10, total_row + 2)


# -----------------------------------------------------------------------------
#  Plan-data-tabel met flight-banners
# -----------------------------------------------------------------------------

def _write_plan_data(ws, plan_df: pd.DataFrame, start_row: int, multi_flight: bool,
                     overlap_factor: float = 0.0) -> int:
    """Schrijf kolom-headers + data + optionele flight-banners + totalen-rij."""
    cols = list(plan_df.columns)

    # Boven de kolom-header een band-header "Verwachte resultaten" over de
    # metric-kolommen (cpm..ecpc). Alleen als die kolommen bestaan.
    metric_names = ["cpm", "cpc", "cpa", "ctr_pct", "gcf",
                    "impressies", "clicks", "bereik", "conversies", "ecpc"]
    metric_cols = [i + 1 for i, c in enumerate(cols) if c in metric_names]
    band_row = start_row
    col_header_row = start_row
    if metric_cols:
        start_c, end_c = min(metric_cols), max(metric_cols)
        # Band-header een rij boven de kolom-headers
        ws.merge_cells(start_row=band_row, start_column=start_c,
                       end_row=band_row, end_column=end_c)
        band_cell = ws.cell(row=band_row, column=start_c, value="Verwachte resultaten")
        band_cell.fill = _FILL_MED; band_cell.font = _BOLD_BLUE; band_cell.border = _BOX
        band_cell.alignment = Alignment(horizontal="center")
        col_header_row = start_row + 1

    # Kolom-headers
    for c, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=col_header_row, column=c, value=col_name)
        cell.fill = _FILL_DARK; cell.font = _BOLD_W; cell.border = _BOX
        cell.alignment = Alignment(horizontal="left")
    row = col_header_row + 1

    # Per-kolom number-format map (openpyxl format-tokens; Excel rendert
    # deze volgens de user-locale, dus '1.447' wordt 1.447 in NL-Excel).
    col_format = {
        "budget":     '"€" #,##0',          # EUR zonder decimals
        "cpm":        '"€" #,##0.00',
        "cpc":        '"€" #,##0.00',
        "cpa":        '"€" #,##0.00',
        "ctr_pct":    '0.0"%"',             # 1.2 -> "1.2%"
        "gcf":        '0.0',
        "impressies": '#,##0',
        "clicks":     '#,##0',
        "bereik":     '#,##0',
        "conversies": '#,##0',
        "ecpc":       '"€" #,##0.00',
    }
    fmt_by_col = {i+1: col_format[c] for i, c in enumerate(cols) if c in col_format}

    # Voor FORMULES: vind kolom-letters voor input-velden, zodat afgeleide
    # cellen (impressies/clicks/bereik/conversies) live mee-updaten in Excel
    # als de planner budget of een kostenmetric aanpast.
    from openpyxl.utils import get_column_letter
    col_idx = {name: i + 1 for i, name in enumerate(cols)}
    letter = {name: get_column_letter(idx) for name, idx in col_idx.items()}
    derived_cols = {"impressies", "clicks", "bereik", "conversies", "ecpc"}
    has_clicks = "clicks" in col_idx
    # Subset die we effectief als formule kunnen schrijven (inputs aanwezig)
    has_budget = "budget" in col_idx
    has_cpm = "cpm" in col_idx
    has_cpc = "cpc" in col_idx
    has_cpa = "cpa" in col_idx
    has_ctr = "ctr_pct" in col_idx
    has_gcf = "gcf" in col_idx
    has_impr = "impressies" in col_idx

    def _derived_formula(col_name, r):
        """Build Excel-formule voor 1 afgeleide kolom op rij r, of None."""
        if col_name == "impressies" and has_budget and has_cpm:
            return f"=IF({letter['cpm']}{r}>0,{letter['budget']}{r}/{letter['cpm']}{r}*1000,\"\")"
        if col_name == "clicks" and has_budget:
            # Voorkeur: cpc. Fallback: impressies * ctr_pct/100.
            cpc_part = f"{letter['budget']}{r}/{letter['cpc']}{r}" if has_cpc else '""'
            ctr_part = (f"{letter['impressies']}{r}*{letter['ctr_pct']}{r}/100"
                        if (has_impr and has_ctr) else '""')
            if has_cpc and has_ctr and has_impr:
                return (f"=IF({letter['cpc']}{r}>0,{cpc_part},"
                        f"IF(AND({letter['ctr_pct']}{r}>0,{letter['impressies']}{r}>0),{ctr_part},\"\"))")
            if has_cpc:
                return f"=IF({letter['cpc']}{r}>0,{cpc_part},\"\")"
            if has_ctr and has_impr:
                return (f"=IF(AND({letter['ctr_pct']}{r}>0,{letter['impressies']}{r}>0),"
                        f"{ctr_part},\"\")")
            return None
        if col_name == "bereik" and has_impr and has_gcf:
            return f"=IF({letter['gcf']}{r}>0,{letter['impressies']}{r}/{letter['gcf']}{r},\"\")"
        if col_name == "conversies" and has_budget and has_cpa:
            return f"=IF({letter['cpa']}{r}>0,{letter['budget']}{r}/{letter['cpa']}{r},\"\")"
        if col_name == "ecpc" and has_budget and has_clicks:
            return f"=IF({letter['clicks']}{r}>0,{letter['budget']}{r}/{letter['clicks']}{r},\"\")"
        return None

    def _write_row(data_row):
        nonlocal row
        for c, col_name in enumerate(cols, start=1):
            if col_name in derived_cols:
                # Afgeleide kolom: altijd formule
                formula = _derived_formula(col_name, row)
                cell = ws.cell(row=row, column=c, value=formula if formula else None)
            elif col_name == "cpc":
                # CPC is half-derived: als planner 'm invult, gebruik waarde.
                # Anders, als clicks bestaan, bereken budget/clicks als formule.
                raw = data_row.get(col_name)
                if raw is not None and not pd.isna(raw) and float(raw) > 0:
                    cell = ws.cell(row=row, column=c, value=float(raw))
                elif has_budget and has_clicks:
                    cell = ws.cell(
                        row=row, column=c,
                        value=f"=IF({letter['clicks']}{row}>0,{letter['budget']}{row}/{letter['clicks']}{row},\"\")",
                    )
                else:
                    cell = ws.cell(row=row, column=c, value=None)
            else:
                val = data_row[col_name]
                if pd.isna(val):
                    val = None
                cell = ws.cell(row=row, column=c, value=val)
            if c in fmt_by_col:
                cell.number_format = fmt_by_col[c]
        row += 1

    if multi_flight and "flight_nr" in cols:
        for flight_nr, group in plan_df.groupby("flight_nr", sort=True):
            naam = str(group["flight_naam"].iloc[0]) if "flight_naam" in group else ""
            fstart = str(group["flight_start"].iloc[0])[:10] if "flight_start" in group else ""
            feind = str(group["flight_eind"].iloc[0])[:10] if "flight_eind" in group else ""
            budget = float(group["budget"].sum()) if "budget" in group else 0.0
            banner = (
                f"Flight {int(flight_nr)} - {naam}    "
                f"{fstart} t/m {feind}    "
                f"media EUR {budget:,.2f}"
            ).replace(",", ".")
            b = ws.cell(row=row, column=1, value=banner)
            b.fill = _FILL_MED; b.font = _BOLD_BLUE; b.border = _BOX
            if len(cols) > 1:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(cols))
            row += 1
            for _, dr in group.iterrows():
                _write_row(dr)
    else:
        for _, dr in plan_df.iterrows():
            _write_row(dr)

    # Totalen-rij na de data
    if len(plan_df) > 0:
        _write_totalen_row(ws, cols, plan_df, total_row=row,
                           overlap_factor=overlap_factor)
        row += 1

    return row


def _write_totalen_row(ws, cols, plan_df: pd.DataFrame, total_row: int,
                       overlap_factor: float) -> None:
    """Schrijf een totalen-rij met **precomputed waarden** (geen formules), zodat
    preview-tools zoals file-explorer de totalen al laten zien. Data-rijen houden
    hun Excel-formules zodat live-update blijft werken. De overlap-factor wordt
    rechtstreeks in het bereik-totaal verwerkt (geen losse cel meer)."""
    col_idx = {name: i + 1 for i, name in enumerate(cols)}

    def _sum(col):
        if col not in plan_df.columns:
            return None
        s = pd.to_numeric(plan_df[col], errors="coerce").fillna(0).sum()
        return float(s) if s else None

    # Sommen uit de plan-data
    budget     = _sum("budget")
    impressies = _sum("impressies")
    clicks     = _sum("clicks")
    conversies = _sum("conversies")
    bereik_raw = _sum("bereik")
    bereik = round(bereik_raw * (1.0 - float(overlap_factor)), 0) if bereik_raw else None

    # Gewogen gemiddeldes
    cpm = round(budget / impressies * 1000, 2) if (budget and impressies) else None
    cpc = round(budget / clicks, 2)            if (budget and clicks)     else None
    cpa = round(budget / conversies, 2)        if (budget and conversies) else None
    ctr = round(clicks / impressies * 100, 2)  if (clicks and impressies) else None
    gcf_net = round(impressies / bereik, 2)    if (impressies and bereik) else None

    totals = {
        "budget": budget,
        "cpm": cpm, "cpc": cpc, "cpa": cpa,
        "ctr_pct": ctr, "gcf": gcf_net,
        "impressies": impressies, "clicks": clicks,
        "bereik": bereik, "conversies": conversies,
        "ecpc": cpc,
    }
    fmt_map = {
        "budget":    '"€" #,##0',
        "cpm":       '"€" #,##0.00', "cpc": '"€" #,##0.00',
        "cpa":       '"€" #,##0.00', "ecpc": '"€" #,##0.00',
        "ctr_pct":   '0.00"%"',      "gcf":  '0.00',
        "impressies":'#,##0',        "clicks": '#,##0',
        "bereik":    '#,##0',        "conversies": '#,##0',
    }

    # Label in kolom A
    lbl = ws.cell(row=total_row, column=1, value="Totaal")
    lbl.fill = _FILL_DARK; lbl.font = _BOLD_W

    # Donkerblauw over hele rij (ook lege cellen) + waarden
    for c, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=total_row, column=c)
        cell.fill = _FILL_DARK
        cell.font = _BOLD_W
        if col_name in totals and totals[col_name] is not None:
            cell.value = totals[col_name]
            if col_name in fmt_map:
                cell.number_format = fmt_map[col_name]

    # Bereik-voetnoot: als er een bereik is, laat rechts in kolom A na totaal-rij een
    # klein subtiel regeltje staan met "Overlap: X%" zodat duidelijk is dat bereik
    # netto is. Geen aparte editable cel meer.
    if bereik is not None and overlap_factor > 0:
        note_row = total_row + 1
        pct = round(overlap_factor * 100)
        note = ws.cell(row=note_row, column=1,
                       value=f"Bereik incl. ontdubbeling: overlap {pct}% toegepast")
        note.font = Font(italic=True, color="808080")


# -----------------------------------------------------------------------------
#  Main writer
# -----------------------------------------------------------------------------

def write_plan_excel(
    pad,
    plan_df: pd.DataFrame,
    flights_df: pd.DataFrame = None,
    budget_df: pd.DataFrame = None,
    ctx: PlanContext = None,
    hide_zero: bool = True,
) -> Path:
    pad = Path(pad)
    pad.parent.mkdir(parents=True, exist_ok=True)

    # Filter 0-rijen
    if hide_zero:
        plan_df = _hide_zero_rows(plan_df, "budget")
        if flights_df is not None:
            flights_df = _hide_zero_rows(flights_df, "budget")
        if budget_df is not None:
            budget_df = _hide_zero_rows(budget_df, "budget_eur")

    multi_flight = _is_multi_flight(plan_df)
    # Single-flight mode: drop flight-kolommen en skip Flights-tab
    if not multi_flight:
        drop = [c for c in ("flight_nr", "flight_naam") if c in plan_df.columns]
        plan_df = plan_df.drop(columns=drop)
        flights_df = None

    # Drop kolommen waar niemand input heeft (voorkomt lege paarse kolommen in Excel)
    def _has_any_value(col):
        if col not in plan_df.columns:
            return False
        s = pd.to_numeric(plan_df[col], errors="coerce").fillna(0)
        return bool((s > 0).any())
    if not _has_any_value("cpa"):
        plan_df = plan_df.drop(columns=[c for c in ("cpa", "conversies") if c in plan_df.columns])
    if not _has_any_value("gcf"):
        plan_df = plan_df.drop(columns=[c for c in ("gcf", "bereik") if c in plan_df.columns])

    # Manueel workbook bouwen (met openpyxl) om banners + header-kader te zetten
    wb = Workbook()
    ws_plan = wb.active
    ws_plan.title = "Plan"

    header = ctx.header if ctx else None
    if header is not None:
        data_start = _write_plan_header(ws_plan, header) + 1
    else:
        data_start = 1
    overlap = float((ctx.overlap_factor if ctx else 0.08) or 0.0)
    _write_plan_data(ws_plan, plan_df, start_row=data_start,
                     multi_flight=multi_flight, overlap_factor=overlap)

    # Vries de rij onder de kolom-headers (band-header + kolom-headers = 2 rijen)
    ws_plan.freeze_panes = ws_plan.cell(row=data_start + 2, column=1).coordinate

    # Overige tabbladen via dataframes
    def _add_sheet(name, df):
        ws = wb.create_sheet(name)
        cols = list(df.columns)
        for c, col_name in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=c, value=col_name)
            cell.fill = _FILL_DARK; cell.font = _BOLD_W; cell.border = _BOX
        for r_idx, (_, r) in enumerate(df.iterrows(), start=2):
            for c, col_name in enumerate(cols, start=1):
                val = r[col_name]
                if pd.isna(val):
                    val = None
                ws.cell(row=r_idx, column=c, value=val)

    if flights_df is not None and len(flights_df):
        _add_sheet("Flights", flights_df)
    if budget_df is not None and len(budget_df):
        _add_sheet("Budget", budget_df)

    if ctx and ctx.benchmarks_used:
        bench_rows = []
        for fase, kanalen in (ctx.benchmarks_used or {}).items():
            if isinstance(kanalen, dict):
                for kanaal, pct in kanalen.items():
                    bench_rows.append({"fase": fase, "kanaal": kanaal, "pct_bench": pct})
        if bench_rows:
            _add_sheet("Benchmarks", pd.DataFrame(bench_rows))

    if ctx and ctx.keybeliefs_used:
        _add_sheet("Keybeliefs", pd.DataFrame(ctx.keybeliefs_used))

    # Meta
    meta_rows = [
        {"veld": "gegenereerd_op", "waarde": datetime.now().isoformat(timespec="seconds")},
        {"veld": "generator_versie", "waarde": (ctx.versie if ctx else "0.3")},
        {"veld": "auteur", "waarde": (ctx.auteur if ctx else "planningsmachine")},
        {"veld": "flight_mode", "waarde": "multi" if multi_flight else "single"},
    ]
    if ctx:
        meta_rows += [
            {"veld": "klant", "waarde": _k(ctx.klant_cfg, "klant", "naam", default="")},
            {"veld": "klant_code", "waarde": str(_k(ctx.klant_cfg, "klant", "code", default="")).upper()},
            {"veld": "campagne", "waarde": ctx.campagne.get("naam", "")},
            {"veld": "campagne_code", "waarde": ctx.campagne.get("code", "")},
            {"veld": "campagne_jaar", "waarde": ctx.campagne.get("jaar", "")},
        ]
        if header:
            meta_rows += [
                {"veld": "totaal_incl_btw", "waarde": header.totaal_incl_btw},
                {"veld": "totaal_ex_btw",   "waarde": header.totaal_ex_btw},
                {"veld": "uren_ex_btw",     "waarde": header.uren_ex_btw},
                {"veld": "media_ex_btw",    "waarde": header.media_ex_btw},
                {"veld": "btw_pct",         "waarde": header.btw_pct},
            ]
    _add_sheet("Meta", pd.DataFrame(meta_rows))
    wb.save(pad)
    return pad
